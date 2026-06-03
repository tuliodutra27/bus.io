#!/usr/bin/env python3
"""
importar_planilha.py — Importa dados reais da planilha para o banco bus.io.

Pré-requisito (instalar uma vez):
    pip install openpyxl

Uso:
    python scripts/importar_planilha.py PLANILHA.xlsx [saida.sql]

    Padrao de saida: dados_reais.sql na pasta corrente.

Para aplicar ao banco (com Docker rodando):
    docker exec -i busio_db sh -c \
        'mysql -u"$MYSQL_USER" -p"$MYSQL_PASSWORD" "$MYSQL_DATABASE"' < dados_reais.sql

Regras de atribuicao de assentos:
  - ADM-04 : usa o numero de assento que esta na planilha.
  - Demais  : assentos atribuidos automaticamente — mulheres primeiro (assentos menores),
              homens em seguida. Para onibus de turno, a atribuicao e feita por letra
              (A/B/C/D), entao diferentes pessoas podem ocupar o mesmo numero de assento
              em letras diferentes.
"""

import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Erro: instale openpyxl antes de rodar:\n  pip install openpyxl")


# ---------------------------------------------------------------------------
# Configuracao
# ---------------------------------------------------------------------------

# (num, aba_adm, aba_turno, cidade)
PARES_ROTA = [
    (1, "01ADM",  "ROTA 01Turno", "Campos dos Goytacazes"),
    (2, "02ADM",  "ROTA 02Turno", "Campos dos Goytacazes"),
    (3, "03ADM",  "ROTA 03Turno", "Campos dos Goytacazes"),
    (4, "04ADM",  "ROTA 04Turno", "Sao Joao da Barra"),
]

# ADM-04 e o unico onibus que usa os assentos da planilha
NUM_ROTA_COM_ASSENTO_PLANILHA = {4}   # somente ADM-04

ADM_CAP_PADRAO = 50
TUR_CAP_PADRAO = 32

LETRAS_VALIDAS = {"A", "B", "C", "D"}

# Primeiros nomes femininos brasileiros (lista abrangente)
NOMES_FEMININOS = {
    "ANA", "ALICE", "ALINE", "ALICIA", "ALBA",
    "AMANDA", "AMELIA", "ADRIANA", "ADRIELE", "ADRIELLA", "ADRIELLY",
    "ALESSANDRA", "ANGELA", "ANNA", "ARIANA",
    "BARBARA", "BEATRIZ", "BRUNA",
    "CAMILA", "CAMILI", "CARLA", "CAROL", "CAROLINA", "CAROLINE",
    "CASSIA", "CARINA", "CINTIA", "CLARICE", "CRISTIANE", "CRISTINA",
    "DANIELA", "DANIELE", "DAIANE", "DEBORA", "DEISE", "DENISE",
    "DIANA", "DIENE",
    "EDILEIA", "EDINA", "EDNA", "EDUARDA",
    "ELAINE", "ELIZABETE", "ELIZANGELA", "ELIANE",
    "ELLEN", "EMILIA", "EMANUELA", "ERICA", "ERIKA", "ERQUILEI",
    "FABIANA", "FABIA", "FERNANDA", "FERNADA", "FLAVIA", "FRANCISCA",
    "GABRIELA", "GABRYELA", "GIOVANA", "GRAZIELA",
    "HERICA", "HERIKA",
    "INGRID", "ISIS", "ISABELA", "ISABELE", "IZIS",
    "JAQUELINE", "JESSICA", "JHENIFER", "JOANA", "JOLANE",
    "JULIANA", "JULIA", "JULLYANA", "JULYA",
    "KISSILA",
    "LAILA", "LARA", "LARISSA", "LAURA", "LAIS", "LEIDYMARA",
    "LIDIA", "LUANA", "LUANE", "LUCIA", "LUCIANA", "LUCIARA",
    "MARCELA", "MARCIA", "MARIANA", "MARISTELA", "MARLENY", "MEIRY",
    "MICHELI", "MICHELLY", "MILENA", "MIRIAN",
    "NADIA", "NATALIA", "NAYARA",
    "ODALIA", "OLIVIA",
    "PAULA", "PAOLA", "PRISCILA", "PRISCILLA",
    "QUESIA",
    "RAFAELA", "RAQUEL", "REBECA", "RENATA", "ROBERTA",
    "ROSA", "ROSANA", "ROSANGELA",
    "SABRINA", "SAMARA", "SANDRA", "SARA", "SARAH", "SIMONE",
    "SOLANGE", "STEPHANY", "SUELEN", "SUELLEN",
    "TAINARA", "TAMARA", "TATIANA", "THAMIRES", "THAYANE",
    "VALDIRENE", "VALERIA", "VANESSA", "VIVIAN", "VIVIANE",
    "YARA",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalizar_nome(txt: str) -> str:
    nfkd = unicodedata.normalize("NFKD", str(txt))
    sem_acento = "".join(c for c in nfkd if not unicodedata.combining(c))
    return " ".join(sem_acento.upper().split())


def detectar_genero(nome: str) -> str:
    primeiro = normalizar_nome(nome).split()[0] if nome else ""
    return "F" if primeiro in NOMES_FEMININOS else "M"


def sql_str(val) -> str:
    if val is None:
        return "NULL"
    return "'" + str(val).replace("\\", "\\\\").replace("'", "\\'") + "'"


def cap_minima(n: int) -> int:
    """Arredonda para o proximo multiplo de 4 >= n."""
    return max(((n + 3) // 4) * 4, 4)


def detectar_colunas(header: tuple) -> dict:
    cols = {}
    for i, v in enumerate(header):
        if v is None:
            continue
        k = normalizar_nome(str(v))
        if "NOME" in k and len(k.split()) <= 3:
            cols.setdefault("nome", i)
        elif "ASSENTO" in k and "N" in k:
            cols.setdefault("assento", i)
        elif k in ("CARGO", "FUNCAO"):
            cols.setdefault("cargo", i)
        elif "HORARIO" in k or "TURNO" in k:
            cols.setdefault("horario", i)
        elif "BAIRRO" in k:
            cols.setdefault("bairro", i)
        elif "TELEFONE" in k or "FONE" in k:
            cols.setdefault("telefone", i)
        elif "LETRA" in k:
            cols.setdefault("letra", i)
        elif "MATRICUL" in k:
            cols.setdefault("matricula", i)
        elif "QUANT" in k:
            cols.setdefault("quant", i)
    return cols


def empresa_regime(horario) -> tuple[str, str]:
    if not horario:
        return ("ALISEO", "turno")
    h = str(horario).strip()
    if "Tercerizadas" in h or "Terceirizadas" in h:
        return ("Terceirizada", "turno")
    if "ADM" in h.upper():
        return ("ALISEO", "admin")
    return ("ALISEO", "turno")


def extrair_nome_rota(ws) -> str:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for v in row1:
        if v and str(v).strip():
            return str(v).strip()
    return ws.title


# ---------------------------------------------------------------------------
# Leitura das abas
# ---------------------------------------------------------------------------

def ler_aba_adm(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    cols = detectar_colunas(rows[1])
    if "nome" not in cols:
        return []

    resultado = []
    for row in rows[2:]:
        nome_val = row[cols["nome"]] if len(row) > cols["nome"] else None
        if not nome_val or not str(nome_val).strip():
            continue

        def cel(key, default=None):
            idx = cols.get(key)
            return row[idx] if idx is not None and idx < len(row) else default

        horario_raw = cel("horario", "Aliseo ADM")
        emp, reg = empresa_regime(horario_raw)

        assento_num = None
        av = cel("assento")
        if av is not None:
            try:
                assento_num = int(av)
            except (ValueError, TypeError):
                pass

        resultado.append({
            "nome":     str(nome_val).strip(),
            "cargo":    str(cel("cargo") or "").strip() or None,
            "bairro":   str(cel("bairro") or "").strip() or None,
            "telefone": None,
            "empresa":  emp,
            "regime":   reg,
            "assento":  assento_num,   # so usado em ADM-04
            "letras":   ["ADM"],
        })
    return resultado


def ler_aba_turno(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    cols = detectar_colunas(rows[1])
    if "nome" not in cols:
        return []

    c_nome = cols["nome"]
    c_ass  = cols.get("assento")
    c_car  = cols.get("cargo")
    c_bai  = cols.get("bairro")
    c_tel  = cols.get("telefone")
    c_hor  = cols.get("horario")
    c_let  = cols.get("letra")
    c_qua  = cols.get("quant")

    def cel(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    resultado = []
    atual = None

    for row in rows[2:]:
        nome_val  = cel(row, c_nome)
        letra_val = cel(row, c_let)
        quant_val = cel(row, c_qua)

        letra = None
        if letra_val:
            ls = str(letra_val).strip().upper()
            if ls in LETRAS_VALIDAS:
                letra = ls

        nome = str(nome_val).strip() if nome_val and str(nome_val).strip() else None

        if nome:
            if atual:
                resultado.append(atual)

            afastado = str(quant_val or "").upper() == "AFASTADO"
            emp, reg = empresa_regime(cel(row, c_hor))

            assento_num = None
            av = cel(row, c_ass)
            if av is not None:
                try:
                    assento_num = int(av)
                except (ValueError, TypeError):
                    pass

            cargo   = str(cel(row, c_car) or "").strip() or None
            bairro  = str(cel(row, c_bai) or "").strip() or None
            tel_raw = cel(row, c_tel)
            telefone = str(tel_raw).strip() if tel_raw else None

            letras = []
            if not afastado and letra:
                letras = [letra]

            atual = {
                "nome":     nome,
                "cargo":    cargo,
                "bairro":   bairro,
                "telefone": telefone,
                "empresa":  emp,
                "regime":   reg,
                "assento":  assento_num,
                "letras":   letras,
                "afastado": afastado,
            }

        elif letra and atual and letra not in atual["letras"]:
            atual["letras"].append(letra)

    if atual:
        resultado.append(atual)

    return resultado


# ---------------------------------------------------------------------------
# Auto-atribuicao de assentos
# ---------------------------------------------------------------------------

def ordenar_por_genero(dados: list[dict]) -> list[dict]:
    """Retorna lista com mulheres primeiro, depois homens (ordem de aparicao preservada)."""
    mulheres = [d for d in dados if detectar_genero(d["nome"]) == "F"]
    homens   = [d for d in dados if detectar_genero(d["nome"]) == "M"]
    return mulheres + homens


def auto_assentos_adm(dados: list[dict]) -> dict[str, int]:
    """
    Auto-atribui assentos para onibus ADM.
    Retorna {nome_norm: num_assento}.
    Mulheres recebem os menores numeros.
    """
    vistos: dict[str, dict] = {}
    for d in dados:
        key = normalizar_nome(d["nome"])
        if key not in vistos:
            vistos[key] = d

    ordenados = ordenar_por_genero(list(vistos.values()))
    return {normalizar_nome(p["nome"]): i + 1 for i, p in enumerate(ordenados)}


def auto_assentos_turno(dados: list[dict]) -> list[tuple[str, int, str]]:
    """
    Auto-atribui assentos para onibus de turno, por letra.
    Mulheres recebem assentos menores dentro de cada letra.
    Retorna lista de (nome_norm, num_assento, letra).
    """
    por_letra: dict[str, list] = {}
    for d in dados:
        for letra in d["letras"]:
            por_letra.setdefault(letra, [])
            key = normalizar_nome(d["nome"])
            if not any(x[0] == key for x in por_letra[letra]):
                por_letra[letra].append((key, d))

    alocacoes = []
    for letra in sorted(por_letra.keys()):
        pessoas = [d for _, d in por_letra[letra]]
        ordenados = ordenar_por_genero(pessoas)
        for i, p in enumerate(ordenados):
            alocacoes.append((normalizar_nome(p["nome"]), i + 1, letra))
    return alocacoes


# ---------------------------------------------------------------------------
# Geracao do SQL
# ---------------------------------------------------------------------------

def gerar_assentos_vals(onibus_id: int, capacidade: int, id_ini: int):
    vals = []
    for n in range(1, capacidade + 1):
        aid   = id_ini + n - 1
        fila  = (n - 1) // 4 + 1
        col   = (n - 1) % 4 + 1
        lado  = "esq" if col <= 2 else "dir"
        vals.append(f"({aid}, {onibus_id}, {n}, {fila}, {col}, '{lado}')")
    return vals, id_ini + capacidade


def gerar_sql(xlsx_path: Path, saida_path: Path):
    print(f"Lendo planilha: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # ------------------------------------------------------------------ #
    # 1. Dados brutos                                                     #
    # ------------------------------------------------------------------ #
    rotas_raw = {}
    for num, aba_adm, aba_turno, cidade in PARES_ROTA:
        adm  = ler_aba_adm(wb[aba_adm])    if aba_adm   in wb.sheetnames else []
        tur  = ler_aba_turno(wb[aba_turno]) if aba_turno in wb.sheetnames else []
        n_adm = extrair_nome_rota(wb[aba_adm])   if aba_adm   in wb.sheetnames else aba_adm
        n_tur = extrair_nome_rota(wb[aba_turno]) if aba_turno in wb.sheetnames else aba_turno
        rotas_raw[num] = {"adm": adm, "turno": tur, "nome_adm": n_adm,
                          "nome_tur": n_tur, "cidade": cidade}
        print(f"  Rota {num}: {len(adm)} ADM, {len(tur)} Turno")
    wb.close()

    # ------------------------------------------------------------------ #
    # 2. Deduplicacao de colaboradores                                    #
    # ------------------------------------------------------------------ #
    colaboradores: dict[str, dict] = {}
    colab_seq = [1]

    def get_colab(d: dict, rota_id: int) -> dict:
        key = normalizar_nome(d["nome"])
        if key not in colaboradores:
            colaboradores[key] = {
                "id":       colab_seq[0],
                "nome":     d["nome"].title(),
                "matricula": f"IMP{colab_seq[0]:05d}",
                "cargo":    d.get("cargo"),
                "bairro":   d.get("bairro"),
                "telefone": d.get("telefone"),
                "empresa":  d.get("empresa", "ALISEO"),
                "regime":   d.get("regime", "turno"),
                "rota_id":  rota_id,
            }
            colab_seq[0] += 1
        return colaboradores[key]

    for num, rd in rotas_raw.items():
        for d in rd["adm"] + rd["turno"]:
            get_colab(d, num)

    # ------------------------------------------------------------------ #
    # 3. Capacidades e mapa de assentos                                   #
    # ------------------------------------------------------------------ #
    adm_cap = {}   # num -> capacidade ADM
    tur_cap = {}   # num -> capacidade Turno

    for num, rd in rotas_raw.items():
        # ADM: auto-expande se necessario
        adm_unique = len({normalizar_nome(d["nome"]) for d in rd["adm"]})
        if num in NUM_ROTA_COM_ASSENTO_PLANILHA:
            # ADM-04: capacidade fixa; assentos vem da planilha
            adm_cap[num] = ADM_CAP_PADRAO
        else:
            adm_cap[num] = cap_minima(max(adm_unique, ADM_CAP_PADRAO))

        # Turno: max pessoas por letra
        por_letra = {}
        for d in rd["turno"]:
            for l in d["letras"]:
                por_letra.setdefault(l, set()).add(normalizar_nome(d["nome"]))
        max_pl = max((len(s) for s in por_letra.values()), default=0)
        tur_cap[num] = cap_minima(max(max_pl, TUR_CAP_PADRAO))

    # Mapa de IDs: onibus + num_assento -> assento_db_id
    assento_map: dict[tuple, int] = {}
    onibus_ids: dict[str, int] = {}  # "adm_N" | "tur_N" -> onibus_id
    oid_seq = [1]
    ass_ptr = [1]

    for num in sorted(rotas_raw):
        for tipo, cap_val in [("adm", adm_cap[num]), ("tur", tur_cap[num])]:
            oid = oid_seq[0]; oid_seq[0] += 1
            onibus_ids[f"{tipo}_{num}"] = oid
            for n in range(1, cap_val + 1):
                assento_map[(oid, n)] = ass_ptr[0]
                ass_ptr[0] += 1

    # ------------------------------------------------------------------ #
    # 4. Monta INSERTs                                                    #
    # ------------------------------------------------------------------ #
    rota_vals   = []
    onibus_vals = []
    ass_vals    = []
    colab_vals  = []
    aloc_vals   = []
    avisos      = []
    aloc_seq    = [1]

    # Rotas e onibus
    for num, rd in sorted(rotas_raw.items()):
        nome_rota = f"{rd['nome_adm']} / {rd['nome_tur']}"
        rota_vals.append(
            f"({num}, {sql_str(nome_rota)}, {sql_str(rd['cidade'])}, NULL, NULL)"
        )
        adm_oid = onibus_ids[f"adm_{num}"]
        tur_oid = onibus_ids[f"tur_{num}"]
        onibus_vals.append(
            f"({adm_oid}, {sql_str(f'ADM-0{num}')}, 'admin', {adm_cap[num]}, {num}, 1, 0)"
        )
        onibus_vals.append(
            f"({tur_oid}, {sql_str(f'TUR-0{num}')}, 'micro', {tur_cap[num]}, {num}, 1, 0)"
        )
        v_adm, _ = gerar_assentos_vals(adm_oid, adm_cap[num], assento_map[(adm_oid, 1)])
        ass_vals += v_adm
        v_tur, _ = gerar_assentos_vals(tur_oid, tur_cap[num], assento_map[(tur_oid, 1)])
        ass_vals += v_tur

    # Colaboradores
    for c in sorted(colaboradores.values(), key=lambda x: x["id"]):
        colab_vals.append(
            f"({c['id']}, {sql_str(c['nome'])}, {sql_str(c['matricula'])}, "
            f"{sql_str(c.get('cargo'))}, {sql_str(c.get('telefone'))}, "
            f"'{c['regime']}', {sql_str('Campos dos Goytacazes')}, "
            f"{sql_str(c.get('bairro'))}, {sql_str(c['empresa'])}, {c['rota_id']})"
        )

    # Alocacoes
    ocupados: dict[tuple, str] = {}   # (onibus_id, assento_num, letra) -> nome

    def add_aloc(onibus_id, num_assento, colab_id, letra, nome_debug, bus_label):
        chave = (onibus_id, num_assento, letra)
        if chave in ocupados:
            avisos.append(
                f"{bus_label}: assento {num_assento} letra {letra} duplicado -- "
                f"{nome_debug} (ja: {ocupados[chave]})"
            )
            return
        cap = adm_cap if "ADM" in bus_label else tur_cap
        num_rota = int(bus_label.split("-0")[1])
        limite = adm_cap[num_rota] if "ADM" in bus_label else tur_cap[num_rota]
        if num_assento < 1 or num_assento > limite:
            avisos.append(f"{bus_label}: assento {num_assento} fora do limite (1-{limite}) -- {nome_debug}")
            return
        ocupados[chave] = nome_debug
        aid = assento_map[(onibus_id, num_assento)]
        aloc_vals.append(f"({aloc_seq[0]}, {aid}, {colab_id}, {sql_str(letra)})")
        aloc_seq[0] += 1

    for num, rd in sorted(rotas_raw.items()):
        adm_oid = onibus_ids[f"adm_{num}"]
        tur_oid = onibus_ids[f"tur_{num}"]

        # ---- ADM ----
        if num in NUM_ROTA_COM_ASSENTO_PLANILHA:
            # ADM-04: usa assentos da planilha
            for d in rd["adm"]:
                c = colaboradores.get(normalizar_nome(d["nome"]))
                if not c or d["assento"] is None:
                    continue
                add_aloc(adm_oid, d["assento"], c["id"], "ADM", d["nome"], f"ADM-0{num}")
        else:
            # Outros ADM: auto-atribui com mulheres primeiro
            mapa = auto_assentos_adm(rd["adm"])
            for nome_norm, num_ass in mapa.items():
                c = colaboradores.get(nome_norm)
                if not c:
                    continue
                add_aloc(adm_oid, num_ass, c["id"], "ADM", c["nome"], f"ADM-0{num}")

        # ---- Turno ----
        # Todos os onibus de turno: auto-atribui por letra com mulheres primeiro
        alocacoes_tur = auto_assentos_turno(rd["turno"])
        for nome_norm, num_ass, letra in alocacoes_tur:
            c = colaboradores.get(nome_norm)
            if not c:
                continue
            add_aloc(tur_oid, num_ass, c["id"], letra, c["nome"], f"TUR-0{num}")

    # ------------------------------------------------------------------ #
    # 5. Arquivo SQL                                                      #
    # ------------------------------------------------------------------ #
    def bloco(tabela, colunas, vals):
        if not vals:
            return []
        return [
            f"INSERT INTO {tabela} ({colunas}) VALUES",
            ",\n".join(f"  {v}" for v in vals) + ";",
            "",
        ]

    linhas = [
        "-- Importacao gerada por scripts/importar_planilha.py",
        "-- Para aplicar:",
        "--   docker exec -i busio_db sh -c",
        "--     'mysql -u\"$MYSQL_USER\" -p\"$MYSQL_PASSWORD\" \"$MYSQL_DATABASE\"'",
        "--     < dados_reais.sql",
        "",
        "SET NAMES utf8mb4;",
        "SET FOREIGN_KEY_CHECKS=0;",
        "TRUNCATE TABLE solicitacoes_hora_extra;",
        "TRUNCATE TABLE excecoes_data;",
        "TRUNCATE TABLE alocacoes_fixas;",
        "TRUNCATE TABLE assentos;",
        "TRUNCATE TABLE colaboradores;",
        "TRUNCATE TABLE onibus;",
        "TRUNCATE TABLE rotas;",
        "DELETE FROM configuracoes WHERE chave = 'turno_letra_ativa';",
        "INSERT INTO configuracoes (chave, valor) VALUES ('turno_letra_ativa', 'A');",
        "SET FOREIGN_KEY_CHECKS=1;",
        "",
    ]
    linhas += bloco("rotas", "id, nome, cidade, bairros, horarios", rota_vals)
    linhas += bloco("onibus", "id, identificador, tipo, capacidade, rota_id, ativo, exemplo", onibus_vals)
    linhas += bloco("assentos", "id, onibus_id, numero, fila, coluna, lado", ass_vals)
    linhas += bloco("colaboradores",
                    "id, nome, matricula, setor, telefone, regime, cidade, bairro, empresa, rota_id",
                    colab_vals)
    linhas += bloco("alocacoes_fixas", "id, assento_id, colaborador_id, turno_letra", aloc_vals)

    saida_path.write_text("\n".join(linhas), encoding="utf-8")

    # Arquivo de avisos separado (sem problemas de encoding no console Windows)
    if avisos:
        avisos_path = saida_path.with_suffix(".avisos.txt")
        avisos_path.write_text("\n".join(avisos), encoding="utf-8")
        print(f"\n  {len(avisos)} aviso(s) gravados em: {avisos_path}")
    else:
        print("\n  Nenhum aviso!")

    # ------------------------------------------------------------------ #
    # 6. Resumo                                                           #
    # ------------------------------------------------------------------ #
    print()
    print("=" * 60)
    print(f"SQL gerado: {saida_path}")
    print(f"  Rotas:         {len(rota_vals)}")
    print(f"  Onibus:        {len(onibus_vals)}")
    print(f"  Assentos:      {len(ass_vals)}")
    print(f"  Colaboradores: {len(colab_vals)}")
    print(f"  Alocacoes:     {len(aloc_vals)}")
    print(f"  Avisos:        {len(avisos)}")
    print("=" * 60)
    print()
    print("Para aplicar (Docker rodando):")
    print("  docker exec -i busio_db sh -c \\")
    print('    \'mysql -u"$MYSQL_USER" -p"$MYSQL_PASSWORD" "$MYSQL_DATABASE"\' \\')
    print("    < dados_reais.sql")
    print()
    print("Depois, no .env do servidor:")
    print("  SEED_ON_START=false")
    print("  docker compose restart app")


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    xlsx = Path(sys.argv[1])
    saida = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("dados_reais.sql")

    if not xlsx.exists():
        sys.exit(f"Arquivo nao encontrado: {xlsx}")

    gerar_sql(xlsx, saida)
